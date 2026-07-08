#!/usr/bin/env python3
"""Brief 96 P6 — assemble the three report tables and export CSV + XLSX.

Table 1 — Isolated MACC: all 22 rows, sorted by £/tCO₂e ascending (Class D / no-saving at
          the bottom).
Table 2 — Cumulative (phasing spine): running capex / EUI / £ / lifetime tCO₂e, vs CRREM
          2026 target 184.1 and the 95 plateau.
Table 3 — EP validation appendix: EP-mappable NZA Δ | EP Δ | Δ% on EUI/heating/cooling.

Run: validation/.venv/bin/python scripts/report/build_tables.py
"""
import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import benchmarks as B  # noqa: E402
from interventions import INTERVENTIONS, PHASING_SPINE, by_ref  # noqa: E402
from metrics import compute_row  # noqa: E402
from offmodel import interlink_1_5, pv_7_1, refrigerant_3_2  # noqa: E402

REPO = Path(__file__).resolve().parents[2]
D = REPO / "docs" / "report"
NZA = json.loads((D / "data/nza_runs.json").read_text())
EP = json.loads((D / "data/ep_runs.json").read_text())
ep_by_ref = {r["ref"]: r for r in EP["rows"]}
CRREM_2026, PLATEAU = 184.1, 95.0

nza_base = NZA["baseline"]
OFF = {"1.5": interlink_1_5(), "7.1": pv_7_1()}


# ── Table 1 — isolated rows for all 22 ────────────────────────────────────────
def row_for(iv):
    ep_ok = ep_by_ref.get(iv["ref"], {}).get("status") in ("done", "cached")
    if iv.get("enabling"):
        return compute_row(iv, ep_validated=False)
    if iv["ref"] in OFF:            # Class C off-model (1.5, 7.1)
        return compute_row(iv, off=OFF[iv["ref"]], ep_validated=False)
    mod = NZA["isolated"].get(iv["ref"])
    if mod is None:                # e.g. 2.3 — no clean static-engine representation
        r = compute_row(iv, ep_validated=False)
        r["flags"] = r["flags"] + ["no-model"]
        return r
    return compute_row(iv, baseline=nza_base, modelled=mod, ep_validated=ep_ok)


rows1 = [row_for(iv) for iv in INTERVENTIONS]


def macc_key(r):
    # ascending £/tCO₂e; measures with no positive saving (None) sink to the bottom.
    v = r["gbp_per_tco2e"]
    return (v is None, v if v is not None else 0)


rows1_sorted = sorted(rows1, key=macc_key)

T1_COLS = ["ref", "name", "theme", "category", "cls", "confidence", "capex_central",
           "capex_low", "capex_high", "annual_elec_kwh", "annual_gas_kwh", "annual_gbp",
           "eui_delta", "lifetime_tco2e", "gbp_per_tco2e", "payback_yrs", "flags",
           "ep_validated", "basis"]


def t1_record(r):
    return {
        "ref": r["ref"], "name": r["name"], "theme": r["theme"], "category": r["category"],
        "class": r["cls"], "confidence": r["confidence"],
        "capex_central_£": r["capex_central"], "capex_low_£": r["capex_low"], "capex_high_£": r["capex_high"],
        "annual_elec_kWh": r["annual_elec_kwh"], "annual_gas_kWh": r["annual_gas_kwh"],
        "annual_£_saving": r["annual_gbp"], "EUI_Δ_kWh_m2": r["eui_delta"],
        "lifetime_tCO2e": r["lifetime_tco2e"], "£_per_tCO2e": r["gbp_per_tco2e"],
        "simple_payback_yr": r["payback_yrs"],
        "flags": ("EP-validated ✓ " if r["ep_validated"] else "") + " ".join(r["flags"]),
        "basis": r["basis"],
    }


# ── Table 2 — cumulative phasing spine ────────────────────────────────────────
cum_states = {s["ref"]: s for s in NZA["cumulative"]["states"]}
cum_order = NZA["cumulative"]["order"]        # modellable refs in spine order
prev_state = nza_base


def baseline_gbp():
    return nza_base["elec_mwh"] * 1000 * B.ELEC_TARIFF_GBP_PER_KWH + nza_base["gas_mwh"] * 1000 * B.GAS_TARIFF_GBP_PER_KWH


run_capex = run_eui = 0.0
run_eui = nza_base["eui"]
run_gbp = 0.0          # cumulative annual £ saving
run_life = 0.0         # cumulative lifetime tCO₂e
prev_cum = nza_base
rows2 = []
for ref in PHASING_SPINE:
    iv = by_ref(ref)
    run_capex += iv["cost"]["central"]
    note = ""
    if iv.get("enabling"):
        note = "enabling (£0 energy)"
    elif ref in OFF:                            # off-model C (1.5, 7.1)
        o = OFF[ref]
        run_eui += o["eui_delta_kwh_m2"]
        run_gbp += o["annual_gbp_saved"]
        run_life += o["lifetime_tco2e"]
        note = "off-model"
    elif ref in cum_states:                     # modellable A/B — marginal from the NZA chain
        st = cum_states[ref]
        marg_elec = (prev_cum["elec_mwh"] - st["elec_mwh"]) * 1000
        marg_gas = (prev_cum["gas_mwh"] - st["gas_mwh"]) * 1000
        life = B.lifetime_carbon_tco2e(marg_elec, marg_gas, iv["life"]) or 0
        if ref == "3.2":
            life += refrigerant_3_2(life_years=iv["life"])["lifetime_tco2e"]
            note = "incl. refrigerant carbon"
        run_eui = st["eui"]
        run_gbp = (baseline_gbp() - (st["elec_mwh"] * 1000 * B.ELEC_TARIFF_GBP_PER_KWH
                                     + st["gas_mwh"] * 1000 * B.GAS_TARIFF_GBP_PER_KWH)) + sum(
            OFF[o]["annual_gbp_saved"] for o in OFF if PHASING_SPINE.index(o) < PHASING_SPINE.index(ref))
        run_life += life
        prev_cum = st
    rows2.append({
        "step": len(rows2) + 1, "ref": ref, "name": iv["name"], "class": iv["cls"],
        "cumulative_capex_£": round(run_capex),
        "cumulative_EUI_kWh_m2": round(run_eui, 1),
        "cumulative_annual_£_saving": round(run_gbp),
        "cumulative_lifetime_tCO2e": round(run_life, 1),
        "EUI_vs_CRREM_2026": round(run_eui - CRREM_2026, 1),
        "EUI_vs_plateau_95": round(run_eui - PLATEAU, 1),
        "note": note,
    })


# ── Table 3 — EP validation appendix ──────────────────────────────────────────
rows3 = []
for r in EP["rows"]:
    if r.get("status") not in ("done", "cached"):
        rows3.append({"ref": r["ref"], "name": r["name"], "status": r.get("status", "failed")})
        continue
    rec = {"ref": r["ref"], "name": by_ref(r["ref"])["name"], "status": r["status"]}
    for k, lbl in (("eui", "EUI"), ("heating_mwh", "heating"), ("cooling_mwh", "cooling")):
        rec[f"{lbl}_NZA_Δ"] = r[k]["nza_delta"]
        rec[f"{lbl}_EP_Δ"] = r[k]["ep_delta"]
        rec[f"{lbl}_Δ%"] = r[k]["delta_pct"]
    rows3.append(rec)


# ── write CSV (Table 1) + XLSX (all three sheets) ─────────────────────────────
t1_recs = [t1_record(r) for r in rows1_sorted]
with open(D / "HIEX_intervention_metrics.csv", "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(t1_recs[0].keys()))
    w.writeheader()
    w.writerows(t1_recs)

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

wb = Workbook()
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(bold=True, color="FFFFFF")


def sheet(title, records, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    if not records:
        return
    cols = list(records[0].keys())
    ws.append(cols)
    for c in ws[1]:
        c.fill = HEAD_FILL; c.font = HEAD_FONT
    for rec in records:
        ws.append(["" if rec.get(c) is None else rec.get(c) for c in cols])
    for i, col in enumerate(cols, 1):
        width = max(len(str(col)), *(len(str(rec.get(col, ""))) for rec in records))
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(width + 2, 8), 60)
    ws.freeze_panes = "A2"


sheet("1 Isolated (MACC)", t1_recs, first=True)
sheet("2 Cumulative (spine)", rows2)
sheet("3 EP validation", rows3)
wb.save(D / "HIEX_intervention_metrics.xlsx")

print(f"[tables] wrote CSV ({len(t1_recs)} rows) + XLSX (3 sheets)")
print(f"[tables] cumulative final: EUI {rows2[-1]['cumulative_EUI_kWh_m2']}, "
      f"capex £{rows2[-1]['cumulative_capex_£']:,}, "
      f"£ saving/yr {rows2[-1]['cumulative_annual_£_saving']:,}, "
      f"lifetime {rows2[-1]['cumulative_lifetime_tCO2e']} tCO₂e")
# best MACC (lowest positive £/tCO₂e)
best = next(r for r in rows1_sorted if r["gbp_per_tco2e"] is not None)
print(f"[tables] best £/tCO₂e: {best['ref']} {best['name'][:30]} = £{best['gbp_per_tco2e']}/tCO₂e")
